from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from copy import copy

def copy_worksheet(source_wb, target_wb, sheet_name):
    if sheet_name not in source_wb.sheetnames:
        print(f"The sheet {sheet_name} does not exist in the source workbook.")
        return

    source_ws = source_wb[sheet_name]

    # If the target sheet already exists in target workbook, you can either overwrite it or create a new one
    if sheet_name in target_wb.sheetnames:
        # Delete the existing sheet
        target_wb.remove(target_wb[sheet_name])
    target_ws = target_wb.create_sheet(title=sheet_name)
    
    # Copying the cell values and styles
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # Copying merged cells
    for range_ in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(range_))

    # Set the column widths to match the source worksheet
    for col in source_ws.columns:
        max_length = 50
        column = col[0].column  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        target_ws.column_dimensions[get_column_letter(column)].width = adjusted_width
    