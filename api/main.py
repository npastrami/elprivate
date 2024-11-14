from quart import Quart, request, jsonify, Response, send_file
from quart_cors import cors
from werkzeug.utils import secure_filename
from dap.azure.uploader import Uploader
from dap.azure.extractor import Extractor
from dap.azure.catcher import Catcher
from azure.storage.blob.aio import BlobServiceClient
import dap.credentials as credentials 
import aiofiles
from io import BytesIO
from dap.csv.copyfunc import copy_worksheet
from database.dapdatabase import DapDatabase
from database.accountdatabase import AccountDatabase
from dap.csv.FOFexport import process_FOF
from dap.azure.sorter import Sorter
from dap.form_mapping_utils import upload_bucket_mapping
from database.ui_table_builder import TableBuilder
import xml.etree.ElementTree as ET
import re
import asyncio
import aiohttp
from pdf2image import convert_from_path
import openpyxl
import traceback
import io
import tempfile
from tortoise import Tortoise, run_async
from tortoise.transactions import in_transaction
from accounts.models.index import TORTOISE_ORM
import accounts.models.index as modelIndex
from accounts.models.role_model import Role
from accounts.routes.auth_routes import auth_routes
from accounts.routes.user_routes import user_routes
from accounts.controllers.user_controller import user_controller
from accounts.controllers.auth_controller import auth_controller
import json
from dap.netchb.entry_xml_builder import EntryXMLBuilder
from accounts.middleware.auth_jwt import verify_token, role_required

app = Quart(__name__)
# Enable CORS for all routes and origins
app = cors(app, allow_origin="http://localhost:8081", allow_methods=["GET", "POST", "OPTIONS", "PUT"], allow_headers=["Content-Type", "Authorization", "x-access-token"])
app.register_blueprint(auth_routes)
app.register_blueprint(user_routes)
app.register_blueprint(user_controller, url_prefix='/api', allow_methods=["GET", "POST", "OPTIONS"], allow_headers=["Content-Type", "Authorization"])
app.register_blueprint(auth_controller, url_prefix='/api/auth', allow_methods=["GET", "POST", "OPTIONS"], allow_headers=["Content-Type", "Authorization"])

db_instance = None

@app.before_serving
async def init():
    global db_instance
    print("starting init")
    await Tortoise.init(config=modelIndex.TORTOISE_ORM)
    await Tortoise.generate_schemas(safe=True)
    db_instance = AccountDatabase(user="postgres", password="newpassword", database="accountdatabase", host="127.0.0.1")
    await db_instance.create_pool()

    await initial()

async def initial():
    # skip if already exists
    if await Role.exists():
        return
    await Role.create(id=1, name="user")
    await Role.create(id=2, name="moderator")
    await Role.create(id=3, name="admin")


@app.route('/api/process_doc', methods=['POST'])
async def process_doc():
    form = await request.form
    client_id = form['clientID']
    version_id = form['versionID']
    form_types = form.getlist('formTypes[]')
    uploaded_files = (await request.files).getlist('files[]')

    tasks = []

    for uploaded_file, form_type in zip(uploaded_files, form_types):
        tasks.append(handle_file(client_id, uploaded_file, form_type, version_id))

    responses = await asyncio.gather(*tasks)

    return jsonify(responses)

async def handle_file(client_id, uploaded_file, form_type, version_id):
    filename = secure_filename(uploaded_file.filename)
    blob_url = await upload_file(client_id, uploaded_file, form_type, version_id)

    if not blob_url:
        return {"status": "Error", "error": "Upload failed"}

    if form_type != 'None':
        xml_str = await extract_data(client_id, filename, upload_bucket_mapping[form_type], form_type, version_id)
        if xml_str:
            return {"status": "Awaiting Review", "xml": xml_str}
        else:
            return {"status": "Empty Extraction"}
    else:
        return {"status": "Upload Completed", "uploaded_file": blob_url}

async def upload_file(client_id, uploaded_file, form_type, version_id):
    bucket_name = upload_bucket_mapping.get(form_type, 'unsorted')
    uploader = Uploader(bucket_name)
    blob_url = await uploader.upload(client_id, uploaded_file)
    if blob_url:
        database = DapDatabase(client_id, blob_url)
        await database.post2postgres_upload(client_id, blob_url, 'uploaded', form_type, bucket_name, version_id)
    return blob_url

async def extract_data(client_id, filename, bucket_name, form_type, version_id):
    sanitized_blob_name = sanitize_blob_name(filename)
    extractor = Extractor(bucket_name)
    extracted_values, blob_sas_url = await extractor.extract(client_id, sanitized_blob_name, form_type)
    await extractor.update_database(client_id, blob_sas_url, filename, form_type, extracted_values, version_id)
    root = ET.Element("W2s")
    for extracted_value in extracted_values:
        w2_element = ET.SubElement(root, "W2")
        for key, value in extracted_value.items():
            if key == 'confidence':
                w2_element.set('confidence', str(value))
            else:
                ET.SubElement(w2_element, key).text = str(value) if value is not None else None
    xml_str = ET.tostring(root, encoding='utf-8').decode('utf-8')
    
    return ({"message": "W2 extraction successful", "xml": xml_str})

def sanitize_blob_name(blob_name):
    # Your existing sanitize function
    sanitized = re.sub(r'[()\[\]{}]', '', blob_name)
    sanitized = sanitized.replace(' ', '_')
    return sanitized

@app.route('/api/get_documents_for_review', methods=['POST'])
async def get_documents_for_review():
    data = await request.json
    client_id = data.get('client_id', None)
    
    db = DapDatabase(client_id, None)
    documents = await db.get_documents_for_review(client_id)
    await db.close()
    
    return jsonify({"documents": documents})

@app.route('/api/get_document_image', methods=['POST'])
async def get_document_image():
    data = await request.get_json()
    doc_name = data.get('doc_name')
    client_id = data.get('client_id', None)
    
    db = DapDatabase(client_id, None)
    doc_url = await db.get_document_image(doc_name)

    if not doc_url:
        return jsonify({"error": "Document not found"}), 404

    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(doc_url) as resp:
                print(f"Status: {resp.status}")
                if resp.status == 200:
                    pdf_content = await resp.read()

                    # Write the PDF content to a temporary file
                    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as temp_pdf:
                        temp_pdf.write(pdf_content)
                        temp_pdf_path = temp_pdf.name

                    # Convert the PDF to images (PNG)
                    images = convert_from_path(temp_pdf_path)

                    # Save the first page as a PNG file in a temporary directory
                    temp_png_path = tempfile.mktemp(suffix=".png")
                    images[0].save(temp_png_path, 'PNG')

                    # Return the PNG file as a response
                    return await send_file(temp_png_path, mimetype='image/png')

                else:
                    return jsonify({"error": "Failed to download document"}), 500

    except Exception as e:
        # Log the exception
        print(f"Error while processing the document: {str(e)}")
        print(traceback.format_exc())
        return jsonify({"error": "Internal server error"}), 500
    
@app.route('/api/approve_document', methods=['POST'])
async def approve_document():
    data = await request.get_json()
    doc_name = data.get('doc_name')
    client_id = data.get('client_id')
    approved = data.get('approved')  # Get the approved status from the request
    edited_fields = data.get('edited_fields', [])  # Get the edited fields from the request
    print("edited fields: ", edited_fields)

    if not doc_name or not client_id:
        return jsonify({"error": "Missing doc_name or client_id"}), 400

    db = DapDatabase(None, None)
    new_status = 'reviewed' if approved else 'extracted'

    # Update the status of the document
    await db.update_review_status(doc_name, client_id, new_status)

    # Update the edited fields in the database
    for field in edited_fields:
        field_name = field.get('field_name')
        field_value = field.get('field_value')
        if field_name and field_value is not None:
            await db.update_field_value(client_id, doc_name, field_name, field_value)

    await db.close()

    return jsonify({"status": "success", "new_status": new_status})

@app.route('/api/get_services', methods=['POST'])
async def get_services():
    data = await request.get_json()
    user_id = data.get('user_id')

    async with db_instance.pool.acquire() as connection:
        services = await connection.fetchval('''
            SELECT services FROM accounts WHERE username = $1;
        ''', user_id)
        
        print(services)

        if services:
            return jsonify({"services": services})
        else:
            return jsonify({"services": {}}), 200
        
@app.route('/api/add_service', methods=['POST'])
async def add_service():
    data = await request.get_json()
    username = data.get('user_id')
    service_type = data.get('service_type')
    documents = data.get('documents')
    notes = data.get('notes')

    service_id = await db_instance.add_service(username, service_type, documents, notes)

    if service_id:
        return jsonify({"message": "Service added successfully", "service_id": service_id}), 201
    else:
        return jsonify({"message": "Failed to add service"}), 500
    
@app.route('/api/get_service_projects', methods=['GET'])
@verify_token
@role_required('admin')
async def get_service_projects():
    async with db_instance.pool.acquire() as connection:
        service_projects = await connection.fetch('SELECT * FROM service_projects')
        service_projects_list = [dict(record) for record in service_projects]
        print(f'services projects:{service_projects_list}')
        return jsonify({'service_projects': service_projects_list})

@app.route('/api/get_admins', methods=['GET'])
@verify_token
@role_required('admin')  # Only allow admins to fetch the user list
async def get_admins():
    async with db_instance.pool.acquire() as connection:
        admin = await connection.fetch('''
            SELECT user_id FROM user_roles WHERE user_id LIKE 'X%'
        ''')
        admin_list = [dict(record) for record in admin]
        print(admin_list)
        return jsonify({'admins': admin_list})
   
@app.route('/api/update_service_project', methods=['POST'])
@verify_token
@role_required('admin')
async def update_service_project():
    data = await request.get_json()
    service_id = data.get('service_id')
    status = data.get('status')
    assigned_admin_id = data.get('assigned_admin_id')
    print("service id:", service_id, "status:", status, "assigned_admin_id:", assigned_admin_id)

    async with db_instance.pool.acquire() as connection:
        await connection.execute('''
            UPDATE service_projects
            SET status = $1, assigned_admin_id = $2
            WHERE service_id = $3
        ''', status, assigned_admin_id, service_id)
    return jsonify({'message': 'Service project updated successfully'})

def prune_empty_elements(element):
    """
    Recursively remove elements that have no text content and no non-empty child elements.
    :param element: The root element to start pruning from.
    :return: True if the element should be removed, False otherwise.
    """
    # Flag to track if we need to remove this element
    should_remove = True

    # Check each child element recursively
    for child in list(element):
        # Recursively prune child; if it should not be removed, mark that this element should be kept
        if prune_empty_elements(child):
            element.remove(child)
        else:
            should_remove = False  # Keep this element if at least one child is non-empty

    # If the element has text content or attributes, mark it to be kept
    if element.text and element.text.strip():
        should_remove = False
    elif element.attrib:  # Keep if there are attributes (assumes attributes contain meaningful data)
        should_remove = False

    # Return True if the element should be removed (i.e., it's empty and has no non-empty children)
    return should_remove

        
@app.route('/api/generate_final_docs', methods=['POST'])
async def generate_final_docs():
    try:
        data = await request.get_json()
        client_id = data.get('client_id')
        doc_names = data.get('doc_names', [])
        send_to_catcher = data.get('send_to_catcher', False)

        # if not client_id or not doc_names:
        #     return jsonify({"error": "client_id and doc_names are required."}), 400

        print(f"Generating final docs for client_id: {client_id}, document: {doc_names}")

        # Use mawb_data directly for testing
        mawb_data = {
            # Entry Number Section
            "entry-no": "98397699",
            "filer-code": "ABC",
            "check-sum": "1",
            
            # Header Elements
            "importer-tax-id": "12-1234567AB",
            "tax-id": "100-20-3000",  # Ultimate consignee tax ID
            "cf-4811": "12-3456789",  # Required tax ID format
            "processing-port": "8888",
            "consignee-name": "Nicholas Pastrana",
            "mode-transportation": "10",
            "entry-port": "8888",
            "entry-type": "01",
            "entry-date": "2024-11-09",
            "payment-type": "2",
            "statement-date": "2024-11-09",
            "charges": "100",
            "gross-weight": "200",
            "description": "COMMERCIAL MERCHANDISE",

            # Remote Entry Information
            "preparer-port": "8888",
            "preparer-office-code": "99",
            "remote-exam-port": "8888",
            
            # Required Header Elements
            "bond-type": "08",
            "surety-code": "123",
            "state-destination": "CA",
            "carrier-code": "ABCD",
            "location-of-goods": "C213",
            "paperless-summary-certification": None,  # Fixed typo in element name
            "certify-cargo-release-via-ace": None,
            
            # Missing Docs Code
            "code": "10",  # Valid code from enumeration
            
            # Required dates
            "tariff-calculation-date": "2024-11-09",
            "import-date": "2024-11-09",
            "arrival-date": "2024-11-09",
            # "inbond-date": "2024-11-09",
            
            # Required ports
            "unlading-port": "8888",
            
            # Manifest Information
            "total-manifest-quantity": "100",
            "master-scac": "ABCD",
            "master-bill": "123456789012",
            "house-scac": "ABCD",
            "house-bill": "987654321021",
            "quantity": "100",
            "unit": "PCS",
            
            # Bill of Lading
            "master-scac": "ABCD",
            "master-bill": "123456789012",
            "house-scac": "ABCD",
            "house-bill": "987654321021",
            "quantity": "100",
            "unit": "PCS",
            
            "ace_entities": [{
                "entity-code": "BY",
                "entity-name": "BUYER NAME",
                "entity-information": {
                    "entity-address": {
                        "address-components": [{
                            "component-type": "01",
                            "address-information": "567"
                        }, {
                            "component-type": "02",
                            "address-information": "23rd st"
                        }],
                        "city": "NEW YORK",
                        "state-province-code": "NY",
                        "postal-code": "10001",
                        "country": "US"
                    }
                }
            }],

            
            # Invoice Information (for invoices/invoice)
            "invoice-no": "12345",
            "aii-supplier-id": "SUPPLIER123",
            "ultimate-consignee-tax-id": "123-45-7894",
            "related-party": "N",  # Empty tag in sample
            "currency-code": "USA",
            "country-origin": "CN",
            
            
            # # Warehouse Entry
            # "warehouse-filer-code": "ABC",
            # "warehouse-entry-no": "1234567",
            # "warehouse-port": "8888",

            # Other Required Elements
            "other-recon": "1",
            "bond-amount": "50000",
            "consolidated-informal-indicator": "P",
            "bond-waiver-reason": "995",
            "entry-date-election-code": "P",
            "presentation-date": "2024-11-09",
            
            # Fee information
            "informal-fee": "0.00",
            "mail-fee": "0.00",
            "manual-surcharge": "0.00",
            
            # CVD information
            "add-cvd-bond-type": "08",
            "add-cvd-stb-amount": "0",
            
            # Boolean flags (as empty elements in XML)
            "electronic-invoice": None,
            "live-entry": None,
            "precalculated": None,
            "transmit": None,
            "via-ace": None,
            
            # Consolidated Entry Information
            "consolidated-entry-no": "99113123",  # Required for consolidated-entry
            "consolidated-entry-filer-code": "ABC",
            # Invoice Information
            "invoice-no": "12345",
            "aii-supplier-id": "SUPPLIER123",
            "ultimate-consignee-tax-id": "123-45-7894",
            "related-party": None,  # Empty element in sample
            
            # Line Item Required Fields
            "export-date": "2024-11-09",
            "country-origin": "CN",
            "manufacturer-id": "CNMANUFACTURER",
            "country-export": "CN",  # Required for ocean/air
            "lading-port": "12345",  # Required for ocean
            "gross-weight": "200",
            
            # Tariff Information
            "tariff": [{
                "tariff-no": "1234567890",
                "value": "20",
                "quantity1": "10.51",
                "unit-of-measure1": "KG",
                "special-program": "A",
                "duty": "2.25"
            },
            {
                "tariff-no": "987654321",
                "value": "30",
                "duty": "3.50",
                "quantity1": "1000000000",
                "unit-of-measure1": "M",
                "quantity2": "50",
                "unit-of-measure2": "CM"
            }],
            
            # Fees Information
            "fees": [{
                "class-code": "056",
                "tariff-no": "1234567890",
                "amount": "100.45"
            }],
            
            # Previous entries remain same...
            
            # Party Information
            "delivered-to": "DELIVERED",
            "sold-to": "SOME SELLER", 
            "exporter": "SOME MFG", 
        }

        # Create an instance of EntryXMLBuilder and build the XML
        mapping_file = "NetCHB_Mapping.xlsx"  # Ensure this path is correct
        ams_builder = EntryXMLBuilder(client_id, mawb_data, mapping_file, root_tag="entry", namespaces={'entry': 'http://www.netchb.com/xml/entry'})
        
        try:
            xml_tree = ams_builder.build_xml_from_mapping()
            prune_empty_elements(xml_tree.root)  # Prune empty elements from the tree
            xml_str = xml_tree.to_string()

            print(f"Generated XML for test:\n{xml_tree.to_pretty_string()}")
        except Exception as e:
            error_message = f"Error during XML generation: {e}"
            print(error_message)
            return jsonify({"error": error_message}), 500

        # Optionally push the XML to sandbox API
        try:
            response = await ams_builder.push_to_netchb(xml_str)
            print(f"Response from sandbox API: {response}")
        except Exception as e:
            error_message = f"Error during XML submission to NetChb API: {e}"
            print(error_message)
            return jsonify({"error": error_message}), 500

        return jsonify({"message": "XML generated and submitted successfully"})

    except Exception as e:
        error_message = f"Error in generate_final_docs: {e}"
        print(error_message)
        return jsonify({"error": error_message}), 500



# catcher = Catcher('approved-docs')
#         db = DapDatabase(client_id, None)
#         final_urls = []

#         approved_docs = await db.get_documents_for_review(client_id)
#         if not approved_docs:
#             return jsonify({"error": "No approved documents found for this client."}), 404

#         for doc_name in doc_names:
#             matching_docs = [doc for doc in approved_docs if doc['doc_name'] == doc_name]

#             if matching_docs:
#                 field_data = {
#                     doc['field_name']: {'value': doc['field_value'], 'confidence': doc['confidence']}
#                     for doc in matching_docs
#                 }

#                 if send_to_catcher:
#                     print("Sending to catcher")
#                     final_url = await catcher.upload_final_document(client_id, doc_name, field_data)
#                     print(f"Final URL for {doc_name}: {final_url}")
#                     final_urls.append(final_url)
#                 else:
#                     # Save to approved_docs database without uploading to Catcher
#                     print("Skipping the catcher")
#                     final_url = None
#                     last_inserted_id = await db.save_approved_document(client_id, doc_name, final_url, field_data)
#                     print(f"Document {doc_name} saved to approved_docs with ID: {last_inserted_id}")
#                     final_urls.append(f"Saved to database with ID: {last_inserted_id}")
                    
                    
@app.route('/api/download_csv/<document_id>', methods=['GET', 'POST'])
async def download_csv(document_id):
    print("Entered download_csv function")
    json_data = await request.json
    client_id = json_data['clientID']
    db = DapDatabase(None, None)
    sanitized_doc_id = sanitize_blob_name(document_id)
    print(sanitized_doc_id)  
    csv_content = await db.generate_csv(sanitized_doc_id, client_id)

    await db.close()

    response = Response(csv_content, mimetype='text/csv')
    response.headers["Content-Disposition"] = f"attachment; filename={document_id}.csv" 
    
    return response

@app.route('/api/get_client_data', methods=['POST'])
async def get_client_data():
    json_data = await request.json
    client_id = json_data['clientID']

    # Use async with to ensure proper initialization and cleanup of TableBuilder
    async with TableBuilder() as table_builder:
        client_data = await table_builder.fetch_client_data(client_id)

    return jsonify({"data": client_data})

async def process_sort(file):
    filename = secure_filename(file.filename)
    print(f'Processing file: {filename}')
    file_stream = io.BytesIO(file.read())

    sorter = Sorter()
    result = await sorter.sort(file_stream)
    file_stream.close()

    return {**result, 'file_name': filename}

@app.route('/api/sort', methods=['POST'])
async def sort():
    print("files received")
    files = await request.files
    files_to_sort = files.getlist('files[]')

    # Schedule all file processing tasks to run concurrently
    sorted_files = await asyncio.gather(*(process_sort(file) for file in files_to_sort))
    
    print(f'sorted_files: {sorted_files}')
    return {'sorted_files': sorted_files}

@app.route('/api/download_all_documents', methods=['POST', 'GET'])
async def download_all_documents():
        data = await request.json
        client_id = data['clientID']
        document_names = data['documentNames']

        # Initialize the BlobServiceClient asynchronously
        blob_service_client = BlobServiceClient.from_connection_string(credentials.CONNECTION_STRING)
        container_client = blob_service_client.get_container_client(credentials.BUCKET_NAME_CUSTOMS)

        # Download the existing FOFtest.xlsx
        blob_client = container_client.get_blob_client('FOFtemplate.xlsx')
        fof_test_stream = BytesIO()
        stream_downloader = await blob_client.download_blob()
        await stream_downloader.readinto(fof_test_stream)
        fof_test_stream.seek(0)
        fof_workbook = openpyxl.load_workbook(fof_test_stream)

        db = DapDatabase(None, None)

        # Create a new workbook
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove the default sheet

        copy_worksheet(fof_workbook, workbook, 'Sheet1') 

        for document_name in document_names:
            sanitized_name = sanitize_blob_name(document_name)
            original_data, fof_data = await db.generate_sheet_data(sanitized_name, client_id)  
            
            # Add original sheet
            original_sheet = workbook.create_sheet(title=sanitized_name)
            for row in original_data:
                original_sheet.append(row)
            
            # Add FOF sheet
            fof_sheet = workbook.create_sheet(title=f"FOF_{sanitized_name}")
            for row in fof_data:
                fof_sheet.append(row)

        fof_sheet_objects = [workbook[sheet_name] for sheet_name in workbook.sheetnames if 'FOF_' in sheet_name]
        
        process_FOF(workbook, fof_sheet_objects) 

        # Save the workbook to a BytesIO object
        output_stream = BytesIO()
        workbook.save(output_stream)
        output_stream.seek(0)

        # Save the workbook to a temporary file
        async with aiofiles.tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            await tmp.write(output_stream.getvalue())
            await tmp.flush()
            response = await send_file(tmp.name, as_attachment=True, attachment_filename=f"{client_id}_Customs_Batch_Data.xlsx")
        return response



if __name__ == "__main__":
    app.run(debug=True, port=8080)
    
    #main branch